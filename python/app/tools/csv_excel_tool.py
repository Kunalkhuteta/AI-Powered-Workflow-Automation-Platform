
"""
CSV/Excel Tool - File Processing
Parse and generate CSV and Excel files
"""

from app.tools.base import BaseTool
from typing import Any, Dict, List, Optional
import csv
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import base64
from datetime import datetime


class CSVExcelTool(BaseTool):
    """
    CSV and Excel file processing tool
    
    Use Cases:
    - Parse CSV files
    - Read Excel spreadsheets
    - Generate CSV files
    - Create Excel files
    - Data extraction
    - Report generation
    - Data transformation
    
    Config:
        operation: Operation type ('read_csv', 'write_csv', 'read_excel', 'write_excel')
        file_content: File content (base64 or text)
        file_path: File path (for local files)
        data: Data to write (list of dicts)
        headers: Column headers (optional)
        delimiter: CSV delimiter (default: ',')
        sheet_name: Excel sheet name (default: 'Sheet1')
        has_headers: Whether CSV has headers (default: true)
    """
    
    def __init__(self):
        super().__init__()
        self.name = "csv_excel"
        self.description = "Process CSV and Excel files"
    
    def execute(self, node_id: str, config: Dict[str, Any], inputs: Dict[str, Any], context: Any) -> Dict[str, Any]:
        """
        Execute CSV/Excel operation
        
        Returns:
            Parsed data or file information
        """
        try:
            operation = config.get('operation', 'read_csv')
            
            # Execute operation
            if operation == 'read_csv':
                result = self._read_csv(config, inputs)
            elif operation == 'write_csv':
                result = self._write_csv(config, inputs)
            elif operation == 'read_excel':
                result = self._read_excel(config, inputs)
            elif operation == 'write_excel':
                result = self._write_excel(config, inputs)
            elif operation == 'parse_csv_text':
                result = self._parse_csv_text(config, inputs)
            else:
                raise ValueError(f"Unknown operation: {operation}")
            
            return result
            
        except Exception as e:
            raise Exception(f"CSV/Excel operation failed: {str(e)}")
    
    def _read_csv(self, config: Dict[str, Any], inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Read and parse CSV file"""
        
        file_content = config.get('file_content')
        file_path = config.get('file_path')
        delimiter = config.get('delimiter', ',')
        has_headers = config.get('has_headers', True)
        
        # Resolve placeholders
        file_content = self._resolve_placeholders(file_content, inputs)
        file_path = self._resolve_placeholders(file_path, inputs)
        
        # Get CSV content
        if file_content:
            # Base64 encoded content
            if isinstance(file_content, str) and file_content.startswith('data:'):
                # Extract base64 part
                csv_text = base64.b64decode(file_content.split(',')[1]).decode('utf-8')
            else:
                csv_text = file_content
        elif file_path:
            # Read from file
            with open(file_path, 'r', encoding='utf-8') as f:
                csv_text = f.read()
        else:
            raise ValueError("Either file_content or file_path is required")
        
        print(f"📄 Parsing CSV with delimiter '{delimiter}'")
        
        # Parse CSV
        reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
        rows = list(reader)
        
        if not rows:
            return {
                'success': True,
                'operation': 'read_csv',
                'rows': [],
                'count': 0
            }
        
        # Convert to list of dicts if has headers
        if has_headers:
            headers = rows[0]
            data_rows = []
            for row in rows[1:]:
                # Pad row if needed
                while len(row) < len(headers):
                    row.append('')
                data_rows.append(dict(zip(headers, row)))
            
            return {
                'success': True,
                'operation': 'read_csv',
                'headers': headers,
                'rows': data_rows,
                'count': len(data_rows),
                'raw_rows': rows
            }
        else:
            return {
                'success': True,
                'operation': 'read_csv',
                'rows': rows,
                'count': len(rows)
            }
    
    def _write_csv(self, config: Dict[str, Any], inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Generate CSV file"""
        
        data = config.get('data', [])
        headers = config.get('headers')
        delimiter = config.get('delimiter', ',')
        
        # Resolve placeholders
        data = self._resolve_placeholders(data, inputs)
        
        if not data:
            raise ValueError("Data is required for write_csv operation")
        
        print(f"📝 Generating CSV with {len(data)} rows")
        
        # Create CSV content
        output = io.StringIO()
        
        # Determine headers
        if isinstance(data, list) and data and isinstance(data[0], dict):
            if not headers:
                headers = list(data[0].keys())
            
            writer = csv.DictWriter(output, fieldnames=headers, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)
        else:
            # List of lists
            writer = csv.writer(output, delimiter=delimiter)
            if headers:
                writer.writerow(headers)
            writer.writerows(data)
        
        csv_content = output.getvalue()
        
        # Encode as base64
        csv_base64 = base64.b64encode(csv_content.encode('utf-8')).decode('utf-8')
        
        return {
            'success': True,
            'operation': 'write_csv',
            'content': csv_content,
            'base64': csv_base64,
            'row_count': len(data),
            'size_bytes': len(csv_content),
            'download_url': f"data:text/csv;base64,{csv_base64}"
        }
    
    def _read_excel(self, config: Dict[str, Any], inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Read Excel file"""
        
        file_content = config.get('file_content')
        file_path = config.get('file_path')
        sheet_name = config.get('sheet_name', 0)  # 0 = first sheet
        has_headers = config.get('has_headers', True)
        
        # Resolve placeholders
        file_content = self._resolve_placeholders(file_content, inputs)
        file_path = self._resolve_placeholders(file_path, inputs)
        
        # Get workbook
        if file_content:
            # Base64 encoded content
            if isinstance(file_content, str) and 'base64,' in file_content:
                excel_bytes = base64.b64decode(file_content.split('base64,')[1])
            else:
                excel_bytes = base64.b64decode(file_content)
            
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        elif file_path:
            wb = openpyxl.load_workbook(file_path)
        else:
            raise ValueError("Either file_content or file_path is required")
        
        # Get sheet
        if isinstance(sheet_name, int):
            ws = wb.worksheets[sheet_name]
        else:
            ws = wb[sheet_name]
        
        print(f"📊 Reading Excel sheet: {ws.title}")
        
        # Read data
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        
        # Remove empty rows
        rows = [row for row in rows if any(cell is not None for cell in row)]
        
        if not rows:
            return {
                'success': True,
                'operation': 'read_excel',
                'sheet_name': ws.title,
                'rows': [],
                'count': 0
            }
        
        # Convert to list of dicts if has headers
        if has_headers:
            headers = [str(h) if h is not None else f'Column{i}' for i, h in enumerate(rows[0])]
            data_rows = []
            
            for row in rows[1:]:
                # Pad row if needed
                while len(row) < len(headers):
                    row.append(None)
                
                # Convert None to empty string
                row = ['' if cell is None else cell for cell in row]
                
                data_rows.append(dict(zip(headers, row)))
            
            return {
                'success': True,
                'operation': 'read_excel',
                'sheet_name': ws.title,
                'headers': headers,
                'rows': data_rows,
                'count': len(data_rows),
                'raw_rows': rows
            }
        else:
            return {
                'success': True,
                'operation': 'read_excel',
                'sheet_name': ws.title,
                'rows': rows,
                'count': len(rows)
            }
    
    def _write_excel(self, config: Dict[str, Any], inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Generate Excel file"""
        
        data = config.get('data', [])
        headers = config.get('headers')
        sheet_name = config.get('sheet_name', 'Sheet1')
        style_headers = config.get('style_headers', True)
        
        # Resolve placeholders
        data = self._resolve_placeholders(data, inputs)
        
        if not data:
            raise ValueError("Data is required for write_excel operation")
        
        print(f"📊 Generating Excel file with {len(data)} rows")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Determine headers
        if isinstance(data, list) and data and isinstance(data[0], dict):
            if not headers:
                headers = list(data[0].keys())
            
            # Write headers
            ws.append(headers)
            
            # Style headers
            if style_headers:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Write data
            for item in data:
                row = [item.get(h, '') for h in headers]
                ws.append(row)
        else:
            # List of lists
            if headers:
                ws.append(headers)
                if style_headers:
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            
            for row in data:
                ws.append(row)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()
        
        # Encode as base64
        excel_base64 = base64.b64encode(excel_bytes).decode('utf-8')
        
        return {
            'success': True,
            'operation': 'write_excel',
            'base64': excel_base64,
            'row_count': len(data),
            'sheet_name': sheet_name,
            'size_bytes': len(excel_bytes),
            'download_url': f"data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}"
        }
    
    def _parse_csv_text(self, config: Dict[str, Any], inputs: Dict[str, Any]) -> Dict[str, Any]:
        """Parse CSV text from string (for inline data)"""
        
        csv_text = config.get('csv_text', '')
        delimiter = config.get('delimiter', ',')
        has_headers = config.get('has_headers', True)
        
        # Resolve placeholders
        csv_text = self._resolve_placeholders(csv_text, inputs)
        
        if not csv_text:
            raise ValueError("csv_text is required")
        
        # Parse
        reader = csv.reader(io.StringIO(csv_text), delimiter=delimiter)
        rows = list(reader)
        
        if has_headers and rows:
            headers = rows[0]
            data_rows = [dict(zip(headers, row)) for row in rows[1:]]
            
            return {
                'success': True,
                'operation': 'parse_csv_text',
                'headers': headers,
                'rows': data_rows,
                'count': len(data_rows)
            }
        else:
            return {
                'success': True,
                'operation': 'parse_csv_text',
                'rows': rows,
                'count': len(rows)
            }
    
    def _resolve_placeholders(self, data: Any, inputs: Dict[str, Any]) -> Any:
        """Resolve {{placeholder}} syntax"""
        if isinstance(data, dict):
            return {k: self._resolve_placeholders(v, inputs) for k, v in data.items()}
        elif isinstance(data, list):
            return [self._resolve_placeholders(item, inputs) for item in data]
        elif isinstance(data, str):
            import re
            pattern = r'\{\{([^}]+)\}\}'
            matches = re.findall(pattern, data)
            
            if not matches:
                return data
            
            if data.strip() == f"{{{{{matches[0]}}}}}":
                return self._get_nested_value(inputs, matches[0].strip())
            
            result = data
            for match in matches:
                value = self._get_nested_value(inputs, match.strip())
                result = result.replace(f"{{{{{match}}}}}", str(value))
            
            return result
        else:
            return data
    
    def _get_nested_value(self, data: Dict, path: str) -> Any:
        """Get nested value using dot notation"""
        keys = path.split('.')
        current = data
        
        for key in keys:
            if isinstance(current, dict):
                current = current.get(key)
            elif isinstance(current, list) and key.isdigit():
                current = current[int(key)]
            else:
                return None
        
        return current


# Export
__all__ = ['CSVExcelTool']